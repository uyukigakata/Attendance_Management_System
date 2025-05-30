import openpyxl
import os
import json
from dotenv import load_dotenv

load_dotenv()

json_path = os.getenv('MEMBER_JDON_PATH', 'members_data.json')

# JSON読み込み
with open(json_path, "r", encoding="utf-8") as f:
    data = json.load(f)

# 新しいExcelファイルを作成
wb = openpyxl.Workbook()
sheet = wb.active

# ヘッダー行を作成
sheet['A1'] = '学籍番号'
sheet['B1'] = '無視する列'
sheet['C1'] = '氏名'

for idx, (student_id, dummy, name) in enumerate(data, start=2):
    sheet.cell(row=idx, column=1, value=student_id)
    sheet.cell(row=idx, column=2, value=dummy)
    sheet.cell(row=idx, column=3, value=name)

# 保存
wb.save('members.xlsx')
print("members.xlsxを作成しました！")
