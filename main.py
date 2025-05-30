import openpyxl as op
import sqlite3
import nfc
import re
import sys
import time
import slackweb
import os
from dotenv import load_dotenv

load_dotenv()


def db_setting(active_sheet, total_row):
    # DBの読み込み
    # 自動コミットモードを設定
    conn = sqlite3.connect("Member.db", isolation_level=None)

    # テーブルを削除するSQL文
    delete_table_sql = """
    DROP TABLE SecMember;
    """
    # テーブルが無い時のエラー処理
    try:
        # SQL文を実行
        conn.execute(delete_table_sql)
    except:
        pass

    # テーブルを作成するSQL文
    create_table_sql = """
    CREATE TABLE SecMember (
        student_id INTEGER(7),
        name TEXT,
        status TEXT
    );
    """
    # SQL文を実行
    conn.execute(create_table_sql)

    for i in range(2, total_row+1):
        student_id = active_sheet.cell(column=1, row=i).value
        name = active_sheet.cell(column=3, row=i).value
        data = (student_id, name, "exit")
        sql = "INSERT INTO SecMember (student_id, name, status) VALUES (?, ?, ?)"
        conn.execute(sql, data)
        
    conn.close()

def show_table():
    # DBの読み込み
    # 自動コミットモードを設定
    conn = sqlite3.connect("Member.db", isolation_level=None)
    c = conn.cursor()
    c.execute("SELECT * FROM SecMember")

    print("-----SecMember Table-----")

    for row in c:
        print(row)
    
    conn.close()
    
def nfc_loading():
    conn = sqlite3.connect("Member.db")     # データベース接続
    slack = slackweb.Slack(url=os.getenv("SLACK_WEBHOOK_URL"))
    clf = nfc.ContactlessFrontend("usb")    # USB接続のNFCリーダーを開く
    prev_time = 0   # 前の学生証タッチ時間
    prev_info = None    # 前の学生証の情報

    # finallyを記述
    try:
        while(True):
            # NFCタグのタッチ認識を開始
            tag = clf.connect(rdwr={'on-connect': lambda tag: False})
            
            # タッチが弱くて読み取れない時のエラー検出
            try:
                nfc_info = tag.dump()   # NFCタグの中身を取り出す
            except nfc.tag.tt3.Type3TagCommandError:
                continue

            the_time = time.time()  # 現在のUnix時間を取得
            dif_time = the_time - prev_time     # 現在と前回のタッチ時間の差を取得
            prev_time = the_time    # 現在のUnix時間を退避させておく

            # カードの情報が前回と同じ ＆ 前回とのタッチ時間の差が3秒より小さい場合, 最初に戻る
            if(nfc_info == prev_info and dif_time < 3):
                prev_info = nfc_info    # 現在のカード情報を退避させておく
                continue

            prev_info = nfc_info    # 現在のカード情報を退避させておく

            # カードが学生証ではなかった場合のエラーを検出
            try:
                _str = nfc_info[4]
                start = _str.index('|') + 2
                end = start + 7
                student_id = [int(_str[start:end])]     # 学籍番号を取得
            except Exception as e:
                continue

            # データベースに学籍番号があるか検索するSQL文
            search_sql = "SELECT name, status FROM SecMember WHERE student_id=?"
            
            # SQL文を実行
            db_info = conn.execute(search_sql, student_id)
            
            # データベースから取得したデータを分解
            for row in db_info:
                name = row[0]   # 氏名を取得
                status = row[1]     # 入退室状況を取得

            try:
                # 入退室状況に応じてステータスを変更する
                if(status == "exit"):
                    slack.notify(text=name + "さんが入室しました")  # Slackに通知
                    update_sql = "UPDATE SecMember SET status='enter' WHERE student_id=?"   # 入退室状況を変更するSQL文
                else:
                    slack.notify(text=name + "さんが退室しました")  # Slackに通知
                    update_sql = "UPDATE SecMember SET status='exit' WHERE student_id=?"    # 入退室状況を変更するSQL文
            except Exception:
                continue

            conn.execute(update_sql, student_id)    # SQL文を実行
            conn.commit()   # SQLの処理を確定

    # 最後に必ず実行される処理
    finally:
        conn.close()    # データベースへの接続を切断
    
def main():
    # Excelデータの読み込み
    wb = op.load_workbook("members.xlsx")
    # シートをアクティブ状態にする
    active_sheet = wb.active
    # Excelの行数を取得
    total_row = active_sheet.max_row
    
    db_setting(active_sheet, total_row)
    show_table()
    
    nfc_loading()

    
if __name__ == "__main__":
    main()
